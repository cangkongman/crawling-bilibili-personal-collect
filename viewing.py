import json
import math
import os
import string

from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor
from openpyxl.drawing.xdr import XDRPositiveSize2D
from openpyxl.styles import Alignment, Font, Border, Side, Color, PatternFill, colors
from openpyxl.utils.units import pixels_to_EMU

'''
插入图片的时候，有两个问题：第一是图片大小问题，第二是图片居中问题。
图片大小问题体现在：
    单位不统一。
    具体什么单位我也不清楚，通过实践得出对应关系：这里 1000 -> excel 26.46cm
    可能不同的电脑有不同的默认单位，所以如果图片大小对不上单元格，就自己试出对应关系，然后在全局变量里改一下
    注: 这里的图片是等比例缩放的，单元格长宽比例不一定符合图片长宽比例，所以不会完全吻合，只会有长或宽的一边吻合
图片居中问题体现在：
    excel图片不能居中，要自己精确设置图片的位置
    我在网上找了一个方法：offset_img
    我自己试了一下，找出对应关系如下：
        7612 - 1像素（单元格）
'''

conversion_offset = 7612
conversion_length = 1000 / 26.46

TwoCell_h = 1.02 * conversion_length

# 下面设置图片的大小，要求是厘米做单位再乘换算关系
# 封面的大小
MaxWidth_cover = 7.62 * conversion_length
MaxHeight_cover = 4 * conversion_length

# 头像的大小
MaxWidth_face = 1.69 * conversion_length
MaxHeight_face = 1.02 * conversion_length

alignment_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
alignment_left_top = Alignment(horizontal="left", vertical="top", wrap_text=True)
alignment_left_bottom = Alignment(horizontal="left", vertical="bottom", wrap_text=True)


def filling(start_loc, end_loc, ws):  # 参数为左上角坐标和右下角坐标，形如'D3','A5'等。ws是worksheet对象。
    x_start = start_loc[0]
    y_start = start_loc[1:len(start_loc)]  # 切片获取坐标的数字部分
    x_end = end_loc[0]
    y_end = end_loc[1:len(end_loc)]
    len_y = int(y_end) - int(y_start) + 1
    alphabet = string.ascii_uppercase  # 导入字母表
    len_x = alphabet.index(x_end) - alphabet.index(x_start) + 1
    # 左上
    temp = start_loc
    ws[temp].border = Border(left=Side(style='thick'), top=Side(style='thick'))
    # 右下
    temp = end_loc
    ws[temp].border = Border(right=Side(style='thick'), bottom=Side(style='thick'))
    # 右上
    temp = x_end + y_start
    ws[temp].border = Border(right=Side(style='thick'), top=Side(style='thick'))
    # 左下
    temp = x_start + y_end
    ws[temp].border = Border(left=Side(style='thick'), bottom=Side(style='thick'))
    # 上
    for i in range(0, len_x - 2):
        temp = alphabet[alphabet.index(x_start) + 1 + i] + y_start
        ws[temp].border = Border(top=Side(style='thick'))
    # 下
    for i in range(0, len_x - 2):
        temp = alphabet[alphabet.index(x_start) + 1 + i] + y_end
        ws[temp].border = Border(bottom=Side(style='thick'))
    # 左
    for i in range(0, len_y - 2):
        temp = x_start + str(int(y_start) + 1 + i)
        ws[temp].border = Border(left=Side(style='thick'))
    # 右
    for i in range(0, len_y - 2):
        temp = x_end + str(int(y_start) + 1 + i)
        ws[temp].border = Border(right=Side(style='thick'))
    return 0


def offset_img(img, row, col):
    p2e = pixels_to_EMU
    h, w = img.height, img.width
    size = XDRPositiveSize2D(p2e(w), p2e(h))
    # 求偏移的长度占整个单元格长度比例

    if col == 0:
        # 封面
        COff = (1 - w / MaxWidth_cover) * 1370000
        if h / MaxHeight_cover > 3 / 4:
            race_r = 4 - h / TwoCell_h
        elif h / MaxHeight_cover > 1 / 2:
            race_r = 3 - h / TwoCell_h
            row += 1
        elif h / MaxHeight_cover > 1 / 4:
            race_r = 2 - h / TwoCell_h
            row += 2
        else:
            race_r = 1 - h / TwoCell_h
            row += 3
    else:
        # 头像
        race_r = 1 - h / TwoCell_h
        COff = (1 - w / MaxWidth_face) * 610000 / 2
    ROff = race_r * 183500
    marker = AnchorMarker(col=col, colOff=COff, row=row, rowOff=ROff)
    img.anchor = OneCellAnchor(_from=marker, ext=size)


def SetTitle(ws, title, i):
    ws.merge_cells(start_row=i, start_column=3, end_row=i + 1, end_column=8)
    ss = 'C' + str(i)
    if len(title) > 30:
        font = Font(name="等线", size=16)
    else:
        font = Font(name="等线", size=20)
    ws[ss] = title
    ws[ss].alignment = alignment_center
    ws[ss].font = font


def SetIntro(ws, intro, i):
    ws.merge_cells(start_row=i + 4, start_column=3, end_row=i + 6, end_column=8)
    ss = 'C' + str(i + 4)
    if len(intro) > 50:
        font = Font(name="等线", size=12)
    else:
        font = Font(name="等线", size=14)
    ws[ss] = intro
    ws[ss].alignment = alignment_left_top
    ws[ss].font = font


def SetCover(ws, img_file, i):
    ws.merge_cells(start_row=i, start_column=1, end_row=i + 7, end_column=2)
    if not os.path.exists(img_file):
        return
    img = Image(img_file)
    if img.width / img.height > MaxWidth_cover / MaxHeight_cover:
        img.height = img.height * MaxWidth_cover / img.width
        img.width = MaxWidth_cover
    else:
        img.width = img.width * MaxHeight_cover / img.height
        img.height = MaxHeight_cover

    offset_img(img, i - 1, 0)
    ws.add_image(img)


def SetFace(ws, img_file, i):
    ws.merge_cells(start_row=i, start_column=9, end_row=i + 1, end_column=9)
    if not os.path.exists(img_file):
        return
    img = Image(img_file)

    if img.width / img.height > MaxWidth_face / MaxHeight_face:
        img.height = img.height * MaxWidth_face / img.width
        img.width = MaxWidth_face
    else:
        img.width = img.width * MaxHeight_face / img.height
        img.height = MaxHeight_face
    offset_img(img, i - 1, 8)
    ws.add_image(img)


def SetNumber(ws, i):
    ws.merge_cells(start_row=i + 4, start_column=9, end_row=i + 6, end_column=9)
    ws['I' + str(i + 4)] = math.ceil(i / 10)
    ws['I' + str(i + 4)].font = Font(name="华文行楷", size=36)
    ws['I' + str(i + 4)].alignment = alignment_center


# 标记已失效视频
def MarkDeleted(ws, i):
    # 不知道为什么填充不了红色，就填充黑色好了，黑底黄字
    ws['I' + str(i + 7)].font = Font(color=colors.YELLOW)
    ws['I' + str(i + 7)].fill = PatternFill(patternType='solid', bgColor=colors.BLACK)
    ws['I' + str(i + 7)] = '已失效'


# 设置播放量、收藏量、弹幕数量、上传时间、发布时间、收藏时间、时长、BV、UPid、UP昵称
def SetSome(ws, value_list, i):
    name_list = ['播放量', '收藏量', '弹幕数量', '上传时间', '发布时间', '收藏时间']
    column_list = ['C', 'D', 'E', 'F', 'G', 'H']
    font = Font(name="等线", size=14)
    num = 0
    for name, column in zip(name_list, column_list):
        ss = column + str(i + 2)
        ws[ss] = name

        ss = column + str(i + 3)
        ws[ss] = value_list[num]
        ws[ss].alignment = alignment_left_bottom
        num += 1

    ws['C' + str(i + 7)] = '时长'
    ws['D' + str(i + 7)] = value_list[6]
    ws['E' + str(i + 7)] = 'BV'
    ws['F' + str(i + 7)] = value_list[7]
    ws['G' + str(i + 7)] = 'UPid'
    ws['H' + str(i + 7)] = value_list[8]

    ws.merge_cells(start_row=i + 2, start_column=9, end_row=i + 3, end_column=9)
    ws['I' + str(i + 2)] = value_list[9]
    ws['I' + str(i + 2)].alignment = alignment_center


def view(RPath, WPath):
    file_name_list = os.listdir(RPath)
    print(file_name_list)
    if os.path.exists(WPath):
        os.remove(WPath)
    wb = Workbook()

    # 遍历每一个文件
    for i in file_name_list:
        ws = wb.create_sheet(i.split('.')[0])
        print(i)

        # 图片文件夹
        Photo_cover_path = '视频封面/' + i.split('.')[0] + '/'
        Photo_face_path = 'up头像/'

        # 改变前8行列宽
        column_list = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']
        for j in column_list:
            ws.column_dimensions[j].width = 20

        count = 1
        with open('收藏夹信息/' + i, 'r', encoding='utf-8') as fp:
            data = json.load(fp)
            for k, v in data.items():
                # 设置边框
                filling('A' + str(count), 'H' + str(count + 7), ws)
                # 设置一些小格子里的信息
                some_list = [
                    v['观众数据']['播放量'],
                    v['观众数据']['收藏量'],
                    v['观众数据']['弹幕数量'],
                    v['三个时间']['上传时间'],
                    v['三个时间']['发布时间'],
                    v['三个时间']['收藏时间'],
                    v['视频信息']['时长'],
                    v['BV'],
                    v['up主']['ID'],
                    v['up主']['昵称']
                ]
                SetSome(ws, some_list, count)
                # 设置标题、简介、封面、头像
                SetTitle(ws, v['视频信息']['标题'], count)
                SetIntro(ws, v['视频信息']['简介'], count)
                SetCover(ws, Photo_cover_path + str(v['BV']) + '.jpg', count)
                SetFace(ws, Photo_face_path + str(v['up主']['ID']) + '.jpg', count)
                # 给每个视频一个编号
                SetNumber(ws, count)
                if v['是否失效']:
                    MarkDeleted(ws, count)
                count += 10
    wb.save(WPath)
    wb.save(WPath)
